import streamlit as st
from io import BytesIO
from openpyxl import load_workbook, Workbook

st.set_page_config(page_title="Excel Splitter", page_icon="📊", layout="centered")
st.title("📊 Excel Splitter — Split Big Files in Half")

st.write("Upload a large `.xlsx` or `.xlsm` file and this app will split it into two Excel files, each with half the rows.")

uploaded = st.file_uploader("Upload Excel File", type=["xlsx", "xlsm"])
sheet_name = st.text_input("Sheet name (optional, defaults to first sheet)", value="")
use_header = st.checkbox("First row is a header", value=True)

def split_workbook(file_bytes, sheet=None, header=True):
    # First pass: count rows
    wb = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]

    # Get header if it exists
    header_row = None
    try:
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    except StopIteration:
        header_row = None

    data_start = 2 if header and header_row else 1
    total = sum(1 for _ in ws.iter_rows(min_row=data_start, values_only=True))
    half = total // 2
    wb.close()

    # Second pass: actually split and write
    wb = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]

    def new_wb():
        w = Workbook(write_only=True)
        s = w.create_sheet("Sheet1")
        if header and header_row:
            s.append(list(header_row))
        return w, s

    wb1, ws1 = new_wb()
    wb2, ws2 = new_wb()

    count = 0
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if count < half:
            ws1.append(list(row))
        else:
            ws2.append(list(row))
        count += 1

    wb.close()

    buf1, buf2 = BytesIO(), BytesIO()
    wb1.save(buf1)
    wb2.save(buf2)
    buf1.seek(0)
    buf2.seek(0)

    return buf1, buf2, total, half, total - half

if uploaded:
    with st.spinner("Splitting your file... this might take a bit ⏳"):
        try:
            buf1, buf2, total, a, b = split_workbook(
                uploaded.read(), sheet=sheet_name or None, header=use_header
            )
            st.success(f"✅ Done! Split {total:,} rows into two files ({a:,} + {b:,}).")

            st.download_button(
                "⬇️ Download Part 1 (.xlsx)", data=buf1, file_name="part1.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            st.download_button(
                "⬇️ Download Part 2 (.xlsx)", data=buf2, file_name="part2.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            st.error(f"⚠️ Error: {e}")

st.caption("Built with 💚 Streamlit + openpyxl. Handles huge files via streaming I/O.")
