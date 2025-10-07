import streamlit as st
from io import BytesIO
from openpyxl import load_workbook, Workbook

st.title("Excel Splitter (Half & Half)")
uploaded = st.file_uploader("Upload .xlsx or .xlsm", type=["xlsx","xlsm"])
sheet_name = st.text_input("Sheet name (optional; default = first sheet)", value="")
use_header = st.checkbox("First row is a header", value=True)

def split_workbook(file_bytes, sheet=None, header=True):
    wb = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]

    # header row
    header_row = None
    it = ws.iter_rows(min_row=1, max_row=1, values_only=True)
    try:
        header_row = next(it)
    except StopIteration:
        header_row = None

    data_start = 2 if header and header_row else 1
    total = sum(1 for _ in ws.iter_rows(min_row=data_start, values_only=True))
    half = total // 2
    wb.close()

    # second pass to write outputs
    wb = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]

    def new_wb():
        w = Workbook(write_only=True)
        s = w.create_sheet("Sheet1")
        if header and header_row:
            s.append(list(header_row))
        return w, s

    wb1, ws1 = new_wb()
    wb2, ws2 = new_wb()

    count = 0
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        (ws1 if count < half else ws2).append(list(row))
        count += 1

    wb.close()

    buf1, buf2 = BytesIO(), BytesIO()
    wb1.save(buf1); wb2.save(buf2)
    buf1.seek(0); buf2.seek(0)
    return buf1, buf2, total, half, total - half

if uploaded:
    try:
        b1, b2, total, a, b = split_workbook(uploaded.read(), sheet=sheet_name or None, header=use_header)
        st.success(f"Split {total} data rows → Part1: {a} rows, Part2: {b} rows.")
        st.download_button("Download Part 1 (.xlsx)", data=b1, file_name="part1.xlsx")
        st.download_button("Download Part 2 (.xlsx)", data=b2, file_name="part2.xlsx")
    except Exception as e:
        st.error(f"Error: {e}")
